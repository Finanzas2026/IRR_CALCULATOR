import streamlit as st
import pandas as pd
import numpy as np
import numpy_financial as npf
import openpyxl
import requests
import io
from io import BytesIO
from datetime import datetime

st.set_page_config(page_title="DCF Project Calculator", layout="wide", page_icon="📊")

st.markdown("""
<style>
section[data-testid="stSidebar"] { display: none; }
.main .block-container { max-width: 1400px; padding-top: 1.5rem; }
.kpi-card {
    background: #ffffff; border-radius: 10px; padding: 14px 10px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.10); text-align: center;
    margin-bottom: 12px; min-height: 110px;
    display: flex; flex-direction: column; justify-content: center; align-items: center;
}
.kpi-label { font-size: 10px; font-weight: 700; color: #666; text-transform: uppercase; letter-spacing: 0.5px; line-height: 1.4; }
.kpi-val   { font-size: 20px; font-weight: 900; color: #0052FF; margin: 6px 0 3px; }
.kpi-sub   { font-size: 10px; color: #999; }
.kpi-val-green { font-size: 20px; font-weight: 900; color: #00875A; margin: 6px 0 3px; }
.section-hdr {
    font-size: 13px; font-weight: 800; color: #0052FF;
    letter-spacing: 2px; text-transform: uppercase;
    border-left: 4px solid #0052FF; padding-left: 10px; margin: 20px 0 10px;
}
.page-title { font-size: 26px; font-weight: 900; color: #0052FF; letter-spacing: 1px; }
.page-sub   { font-size: 13px; color: #888; margin-top: -4px; }
</style>
""", unsafe_allow_html=True)

@st.cache_data(ttl=300)
def _download_xlsx():
    FILE_ID = st.secrets["FILE_ID"]
    r = requests.get(f"https://docs.google.com/spreadsheets/d/{FILE_ID}/export?format=xlsx")
    return r.content

@st.cache_data
def get_projects():
    wb = openpyxl.load_workbook(io.BytesIO(_download_xlsx()), read_only=True)
    return [s for s in wb.sheetnames if s not in ("INSTRUCCIONES", "PLANTILLA")]

@st.cache_data
def load_defaults(project_name: str):
    wb = openpyxl.load_workbook(io.BytesIO(_download_xlsx()), data_only=True)
    ws = wb[project_name]

    header = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    years = []
    for v in header[2:]:
        try:
            y = int(v)
            if 1900 < y < 2200:
                years.append(y)
        except (TypeError, ValueError):
            pass
    n = len(years)

    KNOWN = {"INFLOWS", "OUTFLOWS", "FINANCING"}
    sections = {"INFLOWS": [], "OUTFLOWS": [], "FINANCING": []}
    current = None

    def _f(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    for row in ws.iter_rows(min_row=3, values_only=True):
        sec     = str(row[0]).strip() if row[0] else ""
        concept = str(row[1]).strip() if row[1] else ""

        if sec in KNOWN:
            current = sec
            if concept:
                vals = [_f(v) for v in row[2:2 + n]]
                sections[current].append((concept, vals))
            continue

        if current and concept:
            vals = [_f(v) for v in row[2:2 + n]]
            sections[current].append((concept, vals))

    DEFAULTS = {
        "INFLOWS":   ["Rent", "Sales"],
        "OUTFLOWS":  ["CAPEX", "OPEX", "Rent Comm", "Sales Comm"],
        "FINANCING": ["Debt Draw", "Debt Repay"],
    }

    for sec, default_concepts in DEFAULTS.items():
        loaded = {r[0]: r[1] for r in sections[sec]}
        ordered = []
        for c in default_concepts:
            ordered.append((c, loaded.get(c, [0.0] * n)))
        for c, v in sections[sec]:
            if c not in default_concepts:
                ordered.append((c, v))
        sections[sec] = ordered

    return sections, years

def fmt_usd(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    return f"(${abs(v):,.0f})" if v < 0 else f"${v:,.0f}"

def kpi_card(label, value, sub="", green=False):
    cls = "kpi-val-green" if green else "kpi-val"
    return (f'<div class="kpi-card"><div class="kpi-label">{label}</div>'
            f'<div class="{cls}">{value}</div><div class="kpi-sub">{sub}</div></div>')

CONCEPT_WIDTH = 220

def col_cfg(scols):
    cfg = {"Concepto": st.column_config.TextColumn("Concepto", width=CONCEPT_WIDTH)}
    cfg.update({y: st.column_config.NumberColumn(y, format="$%,.0f", width="small") for y in scols})
    cfg["TOTAL"] = st.column_config.NumberColumn("TOTAL", format="$%,.0f", width="small")
    return cfg

def total_row_style(df, num_cols):
    return df.style.apply(
        lambda r: ["background-color:#1E3A5F;color:white;font-weight:bold"] * len(r), axis=1
    ).format(lambda x: "-" if x == 0 else (f"({abs(x):,.0f})" if x < 0 else f"${x:,.0f}"), subset=num_cols)

def render_section(title, key, section_data, scols, selected):
    st.markdown(f'<div class="section-hdr">{title}</div>', unsafe_allow_html=True)

    labels   = [r[0] for r in section_data]
    n_rows   = len(labels)
    vals_key = f"vals_{key}_{selected}"

    if vals_key not in st.session_state or len(st.session_state[vals_key]) != n_rows:
        st.session_state[vals_key] = [list(r[1]) for r in section_data]

    row_totals = [sum(v) for v in st.session_state[vals_key]]

    df = pd.DataFrame(st.session_state[vals_key], columns=scols)
    df.insert(0, "Concepto", labels)
    df["TOTAL"] = row_totals

    edited = st.data_editor(
        df,
        use_container_width=True,
        num_rows="fixed",
        key=f"editor_{key}_{selected}",
        disabled=["TOTAL"],
        column_config=col_cfg(scols),
        hide_index=True,
    )

    edited[scols] = edited[scols].fillna(0).astype(float)

    result   = []
    new_vals = []
    for i in range(len(edited)):
        concept = str(edited.iloc[i]["Concepto"] or f"Concepto {i+1}")
        vals    = edited.iloc[i][scols].tolist()
        result.append((concept, vals))
        new_vals.append(vals)

    if new_vals != st.session_state[vals_key]:
        st.session_state[vals_key] = new_vals
        st.rerun()

    col_sums  = edited[scols].sum()
    total_val = col_sums.sum()
    total_row = pd.DataFrame([{
        "Concepto": f"▶ TOTAL {key}",
        **col_sums.to_dict(),
        "TOTAL": total_val
    }])

    st.dataframe(
        total_row_style(total_row, scols + ["TOTAL"]),
        use_container_width=True,
        hide_index=True,
        column_config={"Concepto": st.column_config.TextColumn("Concepto", width=CONCEPT_WIDTH)},
    )

    return result

st.markdown('<div class="page-title">📊 DCF PROJECT CALCULATOR</div>', unsafe_allow_html=True)
st.markdown('<div class="page-sub">Selecciona un proyecto · edita las celdas · los resultados se recalculan automáticamente</div>', unsafe_allow_html=True)

projects = get_projects()
col_sel, col_info = st.columns([2, 5])
with col_sel:
    selected = st.selectbox("Proyecto", projects, key="project_selector")
with col_info:
    st.markdown(f"<br><span style='color:#888;font-size:13px'>Fuente: <code>Google Sheets</code> · Hoja: <code>{selected}</code></span>",
                unsafe_allow_html=True)

D, YEARS = load_defaults(selected)
SCOLS = [str(y) for y in YEARS]
N = len(YEARS)
st.divider()

metrics_container = st.container()
st.divider()

inflows   = render_section("INFLOWS — Ingresos",             "INFLOWS",   D["INFLOWS"],   SCOLS, selected)
outflows  = render_section("OUTFLOWS — Costos y Comisiones", "OUTFLOWS",  D["OUTFLOWS"],  SCOLS, selected)
st.caption("Los valores de CAPEX, OPEX y comisiones deben ingresarse como números negativos.")
financing = render_section("FCF FROM FINANCING — Deuda",     "FINANCING", D["FINANCING"], SCOLS, selected)

def sum_by_year(section, n):
    return [sum(vals[i] for _, vals in section) for i in range(n)]

inflows_yr   = sum_by_year(inflows,  N)
outflows_yr  = sum_by_year(outflows, N)
financing_yr = sum_by_year(financing, N)

fcf_no_fin   = [inflows_yr[i] + outflows_yr[i]  for i in range(N)]
fcf_with_fin = [fcf_no_fin[i] + financing_yr[i] for i in range(N)]

def safe_irr(cf):
    try:
        v = npf.irr(cf)
        if v is None:
            return None
        v = float(np.real(v))
        return v if not np.isnan(v) else None
    except Exception:
        return None

irr_no  = safe_irr(fcf_no_fin)
irr_fin = safe_irr(fcf_with_fin)
npv_no  = sum(fcf_no_fin)
npv_fin = sum(fcf_with_fin)

noi_last   = inflows_yr[-1] + outflows_yr[-1]
sales_last = inflows_yr[-1]
cap_rate   = noi_last / sales_last if sales_last != 0 else 0

equity_actual = sum(-fcf_with_fin[i] for i in range(N) if fcf_with_fin[i] < 0)
cash_on_cash  = npv_fin / equity_actual if equity_actual != 0 else None

with metrics_container:
    st.markdown('<div class="section-hdr">INVESTMENT RETURNS — Métricas Clave</div>', unsafe_allow_html=True)
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.markdown(kpi_card("IRR Sin Financiamiento",
                         f"{irr_no*100:.2f}%"  if irr_no  is not None else "—", "Project Closing"), unsafe_allow_html=True)
    k2.markdown(kpi_card("IRR Con Financiamiento",
                         f"{irr_fin*100:.2f}%" if irr_fin is not None else "—", "Project Closing", green=True), unsafe_allow_html=True)
    k3.markdown(kpi_card("NPV Sin Financiamiento", fmt_usd(npv_no),  "Suma FCF"), unsafe_allow_html=True)
    k4.markdown(kpi_card("NPV Con Financiamiento", fmt_usd(npv_fin), "Suma FCF"), unsafe_allow_html=True)
    k5.markdown(kpi_card("Cash-on-Cash",
                         f"{cash_on_cash*100:.2f}%" if cash_on_cash is not None else "—", "NPV / Equity"), unsafe_allow_html=True)

    st.markdown("**Resumen IRR / NPV**")
    st.dataframe(pd.DataFrame({
        "Descripción":    ["IRR Sin Financiamiento", "IRR Con Financiamiento"],
        "TIR":            [f"{irr_no*100:.2f}%"  if irr_no  is not None else "—",
                           f"{irr_fin*100:.2f}%" if irr_fin is not None else "—"],
        "VAN (Suma FCF)": [fmt_usd(npv_no), fmt_usd(npv_fin)],
    }), use_container_width=False, hide_index=True)

st.divider()
st.markdown('<div class="section-hdr">FREE CASH FLOW — Resultados Calculados</div>', unsafe_allow_html=True)

fcf_df = pd.DataFrame(
    {"Concepto": ["FCF (Sin Financiamiento)", "FCF (Con Financiamiento)"]}
    | {str(y): [fcf_no_fin[i], fcf_with_fin[i]] for i, y in enumerate(YEARS)}
    | {"SUBTOTAL": [npv_no, npv_fin]}
)

st.dataframe(
    fcf_df.style.format(
        lambda x: f"({abs(x):,.0f})" if x < 0 else f"${x:,.0f}",
        subset=[str(y) for y in YEARS] + ["SUBTOTAL"],
    ),
    use_container_width=True,
    hide_index=True,
    column_config={"Concepto": st.column_config.TextColumn("Concepto", width=CONCEPT_WIDTH)},
)

st.divider()
st.markdown('<div class="section-hdr">DESCARGAR REPORTE</div>', unsafe_allow_html=True)
fmt_choice = st.radio("Selecciona el formato:", ["Excel (.xlsx)", "PDF (.pdf)"], horizontal=True)

def build_excel():
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        wb_out   = writer.book
        hdr_fmt  = wb_out.add_format({"bold": True, "bg_color": "#0052FF", "font_color": "#FFFFFF", "border": 1, "align": "center"})
        lbl_fmt  = wb_out.add_format({"bold": True, "bg_color": "#EEF2FF", "border": 1, "indent": 1})
        num_fmt  = wb_out.add_format({"num_format": '#,##0;(#,##0)', "border": 1, "align": "right"})
        tot_fmt  = wb_out.add_format({"bold": True, "num_format": '#,##0;(#,##0)', "border": 1, "bg_color": "#DBEAFE", "align": "right"})
        pct_fmt  = wb_out.add_format({"num_format": '0.00%', "border": 1, "align": "right"})
        title_fmt = wb_out.add_format({"bold": True, "font_size": 14, "font_color": "#0052FF"})
        sub_fmt  = wb_out.add_format({"bold": True, "bg_color": "#1E3A5F", "font_color": "#FFFFFF", "border": 1, "indent": 1})
        date_fmt = wb_out.add_format({"italic": True, "font_size": 10, "font_color": "#888888"})

        ws = wb_out.add_worksheet("DCF PROJECT")
        writer.sheets["DCF PROJECT"] = ws
        ws.set_column(0, 0, 32)
        for c in range(1, N + 3):
            ws.set_column(c, c, 14)

        ncols = N + 1
        ts = datetime.now().strftime("%d/%m/%Y  %H:%M")
        ws.merge_range(0, 0, 0, ncols, f"Generado: {ts}", date_fmt)
        ws.merge_range(1, 0, 1, ncols, f"DCF PROJECT — {selected.upper()}   |   Closing", title_fmt)
        ws.write(2, 0, "Concepto", hdr_fmt)
        for c, h in enumerate(SCOLS + ["SUBTOTAL"], 1):
            ws.write(2, c, h, hdr_fmt)

        rn = 3

        def write_sec(title, rows):
            nonlocal rn
            ws.merge_range(rn, 0, rn, ncols, title, sub_fmt)
            rn += 1
            for label, vals in rows:
                ws.write(rn, 0, label, lbl_fmt)
                for c, v in enumerate(vals, 1):
                    ws.write(rn, c, v, num_fmt)
                ws.write(rn, len(vals) + 1, sum(vals), tot_fmt)
                rn += 1
            rn += 1

        write_sec("INFLOWS",            inflows)
        write_sec("OUTFLOWS",           outflows)
        write_sec("FCF FROM FINANCING", financing)
        write_sec("FREE CASH FLOW", [
            ("FCF (Sin Financiamiento)", fcf_no_fin),
            ("FCF (Con Financiamiento)", fcf_with_fin),
        ])

        ws.merge_range(rn, 0, rn, ncols, "INVESTMENT RETURNS", sub_fmt); rn += 1
        for label, val, fmt_ in [
            ("IRR Sin Financiamiento", irr_no  or 0, pct_fmt),
            ("IRR Con Financiamiento", irr_fin or 0, pct_fmt),
            ("NPV Sin Financiamiento", npv_no,       tot_fmt),
            ("NPV Con Financiamiento", npv_fin,      tot_fmt),
        ]:
            ws.write(rn, 0, label, lbl_fmt)
            ws.write(rn, 1, val, fmt_)
            rn += 1

    buf.seek(0)
    return buf.read()

def build_pdf():
    from fpdf import FPDF
    pdf = FPDF(orientation="L", unit="mm", format="A3")
    pdf.add_page()
    pdf.set_margins(10, 10, 10)
    pdf.set_auto_page_break(True, margin=15)

    BLUE  = (0, 82, 255);  LBLUE = (238, 242, 255)
    DARK  = (30, 58, 95);  WHITE = (255, 255, 255); TEXT = (38, 39, 48)

    col_w   = max(18, int(360 / (N + 2)))
    label_w = 50
    hdrs    = SCOLS + ["SUBTOTAL"]

    ts = datetime.now().strftime("%d/%m/%Y  %H:%M")
    pdf.set_font("Helvetica", "I", 8); pdf.set_text_color(136, 136, 136)
    pdf.cell(0, 5, f"Generado: {ts}", ln=True, align="C")
    pdf.set_font("Helvetica", "B", 16); pdf.set_text_color(*BLUE)
    pdf.cell(0, 10, f"DCF PROJECT - {selected.upper()}  |  Closing", ln=True, align="C")
    pdf.set_text_color(*TEXT)
    pdf.ln(3)

    def draw_header():
        pdf.set_fill_color(*BLUE); pdf.set_text_color(*WHITE); pdf.set_font("Helvetica", "B", 7)
        pdf.cell(label_w, 7, "Concepto", border=1, align="C", fill=True)
        for h in hdrs:
            pdf.cell(col_w, 7, h, border=1, align="C", fill=True)
        pdf.ln()

    def draw_sec_title(t):
        pdf.set_fill_color(*DARK); pdf.set_text_color(*WHITE); pdf.set_font("Helvetica", "B", 8)
        pdf.cell(label_w + col_w * len(hdrs), 6, f"  {t}", border=1, fill=True, ln=True)

    def fc(v):
        return "-" if v == 0 else (f"({abs(v):,.0f})" if v < 0 else f"{v:,.0f}")

    def draw_row(label, vals):
        pdf.set_fill_color(*LBLUE); pdf.set_text_color(*TEXT); pdf.set_font("Helvetica", "", 7)
        pdf.cell(label_w, 6, f"  {label}", border=1, fill=True)
        for v in vals:
            pdf.cell(col_w, 6, fc(v), border=1, align="R")
        pdf.cell(col_w, 6, fc(sum(vals)), border=1, align="R", fill=True)
        pdf.ln()

    def draw_fcf(label, vals, sub):
        pdf.set_fill_color(219, 234, 254); pdf.set_text_color(*TEXT); pdf.set_font("Helvetica", "B", 7)
        pdf.cell(label_w, 6, f"  {label}", border=1, fill=True)
        for v in vals:
            pdf.cell(col_w, 6, fc(v), border=1, align="R")
        pdf.cell(col_w, 6, fc(sub), border=1, align="R", fill=True)
        pdf.ln()

    draw_header()
    draw_sec_title("INFLOWS")
    for lbl, vals in inflows:  draw_row(lbl, vals)
    pdf.ln(2)
    draw_sec_title("OUTFLOWS")
    for lbl, vals in outflows: draw_row(lbl, vals)
    pdf.ln(2)
    draw_sec_title("FCF FROM FINANCING")
    for lbl, vals in financing: draw_row(lbl, vals)
    pdf.ln(2)
    draw_sec_title("FREE CASH FLOW")
    draw_fcf("FCF (Sin Financiamiento)", fcf_no_fin,   npv_no)
    draw_fcf("FCF (Con Financiamiento)", fcf_with_fin, npv_fin)
    pdf.ln(5)

    pdf.set_fill_color(*DARK); pdf.set_text_color(*WHITE); pdf.set_font("Helvetica", "B", 8)
    pdf.cell(130, 7, "  INVESTMENT RETURNS", border=1, fill=True, ln=True)
    for lbl, val, is_total in [
        ("IRR Sin Financiamiento", f"{irr_no*100:.2f}%"  if irr_no  is not None else "-", False),
        ("IRR Con Financiamiento", f"{irr_fin*100:.2f}%" if irr_fin is not None else "-", False),
        ("NPV Sin Financiamiento", fmt_usd(npv_no).replace("—", "-"),  True),
        ("NPV Con Financiamiento", fmt_usd(npv_fin).replace("—", "-"), True),
    ]:
        pdf.set_fill_color(*LBLUE); pdf.set_text_color(*TEXT); pdf.set_font("Helvetica", "B", 7)
        pdf.cell(80, 6, f"  {lbl}", border=1, fill=True)
        if is_total:
            pdf.set_fill_color(219, 234, 254); pdf.set_font("Helvetica", "B", 7)
            pdf.cell(50, 6, val, border=1, align="R", fill=True)
        else:
            pdf.set_fill_color(*WHITE); pdf.set_font("Helvetica", "", 7)
            pdf.cell(50, 6, val, border=1, align="R", fill=True)
        pdf.ln()

    buf = BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf.read()

if fmt_choice == "Excel (.xlsx)":
    st.download_button("⬇️ Descargar Excel", build_excel(),
                       file_name=f"DCF_{selected}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       type="primary")
else:
    st.download_button("⬇️ Descargar PDF", build_pdf(),
                       file_name=f"DCF_{selected}.pdf",
                       mime="application/pdf", type="primary")
